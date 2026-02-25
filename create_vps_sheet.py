#!/usr/bin/env python3
"""Tạo file vps_ip_pass.xlsx mẫu cho 6 VPS đầu (cột API_KEY đã điền sẵn)."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

KEYS_VPS_1_6 = [
    "fb8c30dd6b2e62d26b0bde004c09fe34",
    "b4dee72f6859aa8aaa7c81164a84355c",
    "7206bf65db1d628672b297955b949215",
    "5433e11fe0a526e77b9b1dbc9ad9029d",
    "19cf9d6ee40b2a1056e59c8dc2b364b2",
    "f18cd842423fcb3363b1139b79024c8f",
]

def main():
    path = Path(__file__).parent / "vps_ip_pass.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VPS"
    headers = ["STT", "IP", "User", "Password", "API_KEY"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
        ws.cell(row=1, column=c).font = Font(bold=True)
    for i, key in enumerate(KEYS_VPS_1_6, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value="")   # IP - điền sau
        ws.cell(row=i + 1, column=3, value="root")
        ws.cell(row=i + 1, column=4, value="")   # Password - điền sau
        ws.cell(row=i + 1, column=5, value=key)
    wb.save(path)
    print(f"Đã tạo: {path}")
    print("Bạn chỉ cần điền cột IP và Password (User mặc định root), rồi chạy: python read_vps_and_show_commands.py")

if __name__ == "__main__":
    main()
